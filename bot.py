import os
import logging
import io
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

BOT_TOKEN = os.environ.get("BOT_TOKEN", "")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def transform_price(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active
    
    max_row = ws.max_row
    
    # Сохраняем высоты строк
    original_row_heights = {}
    for row in range(1, max_row + 1):
        if ws.row_dimensions[row].height:
            original_row_heights[row] = ws.row_dimensions[row].height
    
    # Сохраняем заголовки из строки 3
    # После удаления столбцов останутся: 3(фото), 5(наим), 6(код), 7(страна), 11(штук), 12(годен), 22(вес), 23(объём)
    original_headers = {
        1: ws.cell(row=3, column=3).value or "Фото",
        2: ws.cell(row=3, column=5).value or "Наименование", 
        3: ws.cell(row=3, column=6).value or "Код",
        4: ws.cell(row=3, column=7).value or "Страна",
        5: ws.cell(row=3, column=11).value or "Штук в блоке",
        6: ws.cell(row=3, column=12).value or "Годен до",
        7: ws.cell(row=3, column=22).value or "Вес (кг)",
        8: ws.cell(row=3, column=23).value or "Объём (м3)",
        9: "Цены"
    }
    
    # Сохраняем цены
    price_data = {}
    for row in range(1, max_row + 1):
        prices = []
        m_val = ws.cell(row=row, column=13).value
        o_val = ws.cell(row=row, column=15).value
        q_val = ws.cell(row=row, column=17).value
        s_val = ws.cell(row=row, column=19).value
        
        if m_val and m_val != 0 and not isinstance(m_val, str):
            try:
                prices.append(f"250тр: {int(m_val)}₽")
            except:
                pass
        if o_val and o_val != 0 and not isinstance(o_val, str):
            try:
                prices.append(f"100т: {int(o_val)}₽")
            except:
                pass
        if q_val and q_val != 0 and not isinstance(q_val, str):
            try:
                prices.append(f"50т: {int(q_val)}₽")
            except:
                pass
        if s_val and s_val != 0 and not isinstance(s_val, str):
            try:
                prices.append(f"25тр: {int(s_val)}₽")
            except:
                pass
        
        if prices:
            price_data[row] = "\n".join(prices)
    
    # Очищаем путь к фото
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=3)
        if cell.value and "D:\\" in str(cell.value):
            cell.value = None
    
    # Разъединяем ячейки
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))
    
    # Извлекаем и уменьшаем изображения
    new_images = []
    for img in ws._images:
        if hasattr(img.anchor, '_from'):
            anchor = img.anchor._from
            if anchor.col == 2:
                row_num = anchor.row
                try:
                    img_data = img._data()
                    pil_img = Image.open(io.BytesIO(img_data))
                    pil_img.thumbnail((70, 70), Image.LANCZOS)
                    
                    buf = io.BytesIO()
                    pil_img.save(buf, format='PNG')
                    buf.seek(0)
                    
                    new_img = XLImage(buf)
                    new_images.append((row_num, new_img))
                except:
                    pass
    
    ws._images = []
    
    # Удаляем столбцы
    cols_to_delete = [29, 28, 27, 26, 25, 24, 21, 20, 19, 18, 17, 16, 15, 14, 13, 10, 9, 8, 4, 3, 2]
    for col in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col)
    
    # Добавляем цены
    for row, combined_price in price_data.items():
        cell = ws.cell(row=row, column=9)
        if not isinstance(cell, MergedCell):
            cell.value = combined_price
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(size=12)
    
    # Ширины столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 18
    
    # Удаляем первые 6 строк (контакты и старые заголовки)
    ws.delete_rows(1, 6)
    
    # Вставляем новую строку для заголовков
    ws.insert_rows(1)
    
    # Добавляем заголовки
    for col, header in original_headers.items():
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 25
    
    # Добавляем изображения (сдвиг: -6 за удалённые строки, +1 за новый заголовок, +1 за 0-индекс)
    for row_num, img in new_images:
        new_row = row_num - 6 + 1 + 1
        if new_row > 1:
            img.anchor = f'A{new_row}'
            ws.add_image(img)
    
    # Применяем высоты строк
    for old_row, height in original_row_heights.items():
        new_row = old_row - 6 + 1
        if new_row > 1:
            ws.row_dimensions[new_row].height = height
    
    wb.save(output_path)
    return {"success": True, "rows": ws.max_row, "images": len(ws._images)}


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет! Отправь мне Excel файл с прайсом.\n\n"
        "Я обработаю его и верну с фото и ценами."
    )


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    
    if not document.file_name.endswith(('.xlsx', '.xls')):
        await update.message.reply_text("❌ Отправь Excel файл (.xlsx)")
        return
    
    await update.message.reply_text("⏳ Обрабатываю...")
    
    try:
        file = await context.bot.get_file(document.file_id)
        input_path = f"/tmp/input_{document.file_name}"
        output_path = f"/tmp/telegram_{document.file_name}"
        
        await file.download_to_drive(input_path)
        
        result = transform_price(input_path, output_path)
        
        if result["success"]:
            await update.message.reply_document(
                document=open(output_path, 'rb'),
                filename=f"telegram_{document.file_name}",
                caption=f"✅ Готово! Строк: {result['rows']}, Фото: {result['images']}"
            )
        
        if os.path.exists(input_path):
            os.remove(input_path)
        if os.path.exists(output_path):
            os.remove(output_path)
            
    except Exception as e:
        logger.error(f"Error: {e}")
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")


def main():
    if not BOT_TOKEN:
        print("❌ BOT_TOKEN не установлен!")
        return
    
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    print("🤖 Бот запущен!")
    app.run_polling(drop_pending_updates=True)


if __name__ == '__main__':
    main()
